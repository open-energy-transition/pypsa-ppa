from __future__ import annotations

import io

from openpyxl import load_workbook

from ppa.financial_model import run_project_finance
from ppa.financial_model_excel import export_financial_model


def test_export_financial_model_produces_a_valid_workbook(energy_inputs, pf_inputs):
    result = run_project_finance(pf_inputs, energy_inputs)
    data = export_financial_model(pf_inputs, energy_inputs, result)

    assert isinstance(data, bytes)
    wb = load_workbook(io.BytesIO(data))
    expected_sheets = {"Inputs", "Energy", "Model", "Outputs", "Notes"}
    assert expected_sheets.issubset(set(wb.sheetnames))


def test_export_financial_model_includes_hourly_sheets_when_year_results_given(
    solved_result, energy_inputs, pf_inputs
):
    result = run_project_finance(pf_inputs, energy_inputs)
    data = export_financial_model(
        pf_inputs, energy_inputs, result, year_results=[solved_result]
    )

    wb = load_workbook(io.BytesIO(data))
    hourly_sheets = [name for name in wb.sheetnames if name.startswith("Hourly ")]
    assert len(hourly_sheets) >= 1
    # Hourly sheets are ordered last (bulky per-year data behind headline sheets).
    assert (
        wb.sheetnames[-1] in hourly_sheets
        or wb.sheetnames[-len(hourly_sheets) :] == hourly_sheets
    )


def test_export_financial_model_outputs_sheet_has_key_kpis(energy_inputs, pf_inputs):
    result = run_project_finance(pf_inputs, energy_inputs)
    data = export_financial_model(pf_inputs, energy_inputs, result)
    wb = load_workbook(io.BytesIO(data))
    outputs_ws = wb["Outputs"]
    all_text = " ".join(
        str(cell.value)
        for row in outputs_ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    for label in ("IRR", "NPV", "DSCR"):
        assert label in all_text
