"""Export a streamlined, *live* project-finance workbook.

Produces an ``.xlsx`` that mirrors :mod:`ppa.financial_model`: editable Inputs, a
pre-filled Energy (PyPSA interface) sheet, a transposed annual Model sheet and an
Outputs sheet. The revenue → EBITDA → depreciation → tax → cash-flow chain and
the IRR/NPV/DSCR outputs are written as **live Excel formulas**, so an analyst
can change a tariff, cost or rate and watch the returns update. The debt sizing
(front-loaded drawdown, IDC, DSCR tranche split) is circular by nature, so it is
written as toolkit-computed values that the live formulas reference: clearly
flagged so it can be overridden.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ppa.financial_model import (
    ProjectFinanceInputs,
    EnergyInputs,
    ProjectFinanceResult,
    _build_timeline,
)

# ── Styling ──────────────────────────────────────────────────────────────────
_TITLE = Font(bold=True, size=14, color="1F4E78")
_HEADER = Font(bold=True, color="FFFFFF")
_SECTION = Font(bold=True, size=11, color="1F4E78")
_INPUT_FONT = Font(color="0000CC")  # blue = editable input (convention)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_PREFILL = PatternFill("solid", fgColor="E2EFDA")
_SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
_thin = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _pcol(period: int) -> int:
    """Spreadsheet column index for a 1-based model period (period 1 -> col D=4)."""
    return 3 + period


def export_financial_model(
    p: ProjectFinanceInputs,
    e: EnergyInputs,
    result: ProjectFinanceResult,
    year_results: list | None = None,
) -> bytes:
    wb = Workbook()
    inputs_cells = _write_inputs(wb, p)
    # Per-year hourly dispatch sheets; the Energy totals roll up from these.
    hourly_refs = _write_hourly_sheets(wb, year_results) if year_results else None
    energy_cells = _write_energy(wb, e, hourly_refs)
    _write_model(wb, p, e, result, inputs_cells, energy_cells)
    _write_outputs(wb, result)
    _write_notes(wb)

    # Order: headline sheets first, the bulky per-year hourly sheets at the end.
    hourly = [ws for ws in wb._sheets if ws.title.startswith("Hourly ")]
    others = [ws for ws in wb._sheets if not ws.title.startswith("Hourly ")]
    wb._sheets = others + hourly
    wb.active = 0

    # Recalculate formulas on open
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Inputs sheet ──────────────────────────────────────────────────────────────


def _write_inputs(wb: Workbook, p: ProjectFinanceInputs) -> dict[str, str]:
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60

    ws["B1"] = "Financial Model: Inputs"
    ws["B1"].font = _TITLE
    ws["B2"] = (
        "Yellow cells are editable assumptions. Costs in €M/MW (€M/MWh for BESS)."
    )
    ws["B2"].font = Font(italic=True, color="808080")

    cells: dict[str, str] = {}
    row = 4

    def section(title: str) -> None:
        nonlocal row
        ws.cell(row, 2, title).font = _SECTION
        for c in range(2, 6):
            ws.cell(row, c).fill = _SECTION_FILL
        row += 1

    def field(label: str, key: str, value, unit: str = "", note: str = "") -> None:
        nonlocal row
        ws.cell(row, 2, label)
        vc = ws.cell(row, 3, value)
        vc.fill = _INPUT_FILL
        vc.font = _INPUT_FONT
        vc.border = _BORDER
        if isinstance(value, float):
            vc.number_format = "#,##0.0000" if abs(value) < 10 else "#,##0.00"
        ws.cell(row, 4, unit).font = Font(color="808080")
        if note:
            ws.cell(row, 5, note).font = Font(italic=True, color="A0A0A0")
        cells[key] = f"Inputs!$C${row}"
        row += 1

    section("Build cost")
    field("Onshore wind build cost", "onsw_build_cost", p.onsw_build_cost, "€M/MW")
    field("Solar PV build cost", "pv_build_cost", p.pv_build_cost, "€M/MW")
    field("BESS build cost", "bess_build_cost", p.bess_build_cost, "€M/MWh")
    row += 1
    section("Connection cost")
    field(
        "Onshore wind connection",
        "onsw_connection_cost",
        p.onsw_connection_cost,
        "€M/MW",
    )
    field("Solar PV connection", "pv_connection_cost", p.pv_connection_cost, "€M/MW")
    field("BESS connection", "bess_connection_cost", p.bess_connection_cost, "€M/MWh")
    row += 1
    section("Project development cost (devex)")
    field(
        "Onshore wind devex",
        "onsw_devex",
        p.onsw_devex,
        "€M/MW",
        "Paid as a lump sum at FID, 100% equity-funded.",
    )
    field("Solar PV devex", "pv_devex", p.pv_devex, "€M/MW")
    field("BESS devex", "bess_devex", p.bess_devex, "€M/MWh")
    row += 1
    section("Fixed O&M (p.a.)")
    field("Onshore wind fixed O&M", "onsw_fixed_om", p.onsw_fixed_om, "€M/MW")
    field("Solar PV fixed O&M", "pv_fixed_om", p.pv_fixed_om, "€M/MW")
    field("BESS fixed O&M", "bess_fixed_om", p.bess_fixed_om, "€M/MWh")
    field("Ancillary services", "ancillary_pct", p.ancillary_pct, "% of revenue")
    row += 1
    section("Timing (years)")
    field("Model duration", "model_duration", p.model_duration, "years")
    field(
        "FID period",
        "fid_period",
        p.fid_period,
        "period",
        "Devex is paid and construction begins here.",
    )
    field(
        "Onshore wind construction", "onsw_constr_years", p.onsw_constr_years, "years"
    )
    field("Solar PV construction", "pv_constr_years", p.pv_constr_years, "years")
    field("BESS construction", "bess_constr_years", p.bess_constr_years, "years")
    field("Operating life", "operating_life", p.operating_life, "years")
    row += 1
    section("Revenue")
    field("PPA contract tenor", "ppa_tenor", p.ppa_tenor, "years")
    field("PPA tariff (base)", "ppa_tariff", p.ppa_tariff, "€/MWh")
    field("Penalty multiple", "penalty_multiple", p.penalty_multiple, "×")
    field("LGC / GO price", "lgc_price", p.lgc_price, "€/MWh")
    row += 1
    section("Indexation (% p.a.)")
    field(
        "Indexation offset",
        "indexation_offset_years",
        p.indexation_offset_years,
        "years",
    )
    field("Cost inflation", "cost_inflation", p.cost_inflation, "%")
    field("PPA & LGC indexation", "ppa_indexation", p.ppa_indexation, "%")
    field(
        "Solar-hour price inflation",
        "solar_price_inflation",
        p.solar_price_inflation,
        "%",
    )
    field(
        "Non-solar-hour price inflation",
        "nonsolar_price_inflation",
        p.nonsolar_price_inflation,
        "%",
    )
    row += 1
    section("Project finance")
    field("Debt repayment tenor", "debt_tenor", p.debt_tenor, "years")
    field("Debt rate", "debt_rate", p.debt_rate, "%")
    field("DSCR hurdle (contracted)", "dscr_contracted", p.dscr_contracted, "ratio")
    field(
        "DSCR hurdle (uncontracted)", "dscr_uncontracted", p.dscr_uncontracted, "ratio"
    )
    field(
        "Max gearing (contracted)",
        "max_gearing_contracted",
        p.max_gearing_contracted,
        "%",
    )
    field(
        "Max gearing (uncontracted)",
        "max_gearing_uncontracted",
        p.max_gearing_uncontracted,
        "%",
    )
    row += 1
    section("Depreciation & tax")
    field(
        "Book depreciation rate",
        "book_depreciation_rate",
        p.book_depreciation_rate,
        "%",
    )
    field(
        "Tax depreciation rate", "tax_depreciation_rate", p.tax_depreciation_rate, "%"
    )
    field("Corporate tax rate", "corp_tax_rate", p.corp_tax_rate, "%")
    field("Discount rate (WACC)", "discount_rate", p.discount_rate, "%")

    return cells


# ── Hourly dispatch sheets (one per simulated year) ──────────────────────────

# Fixed row layout for the per-sheet annual aggregate block, so the Energy sheet
# can reference these cells by address.
_AGG_ROWS = {
    "scale": 4,
    "ppa_gwh": 5,
    "excess_solar_gwh": 6,
    "excess_nonsolar_gwh": 7,
    "penalty_gwh": 8,
    "total_solar_gwh": 9,
    "total_nonsolar_gwh": 10,
    "sell_solar_price": 11,
    "sell_nonsolar_price": 12,
    "purchase_price": 13,
    "marketbuy_gwh": 14,
}
_HOURLY_HEADER_ROW = 16
_HOURLY_DATA_START = 17

# Hourly data columns (1-based): timestamp, hour, then the energy/price series.
_HOURLY_COLS = [
    "Timestamp",
    "Hour",
    "Wind (MWh)",
    "PV (MWh)",
    "BESS discharge (MWh)",
    "BESS charge (MWh)",
    "Total generation (MWh)",
    "Market buy (MWh)",
    "Market sell (MWh)",
    "PPA delivered (MWh)",
    "Penalty (MWh)",
    "Price (€/MWh)",
]
# Column letters for the formulas below
_C_HOUR, _C_TOTAL, _C_BUY, _C_SELL, _C_PPA, _C_PEN, _C_PRICE = (
    "B",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
)


def _write_hourly_sheets(wb: Workbook, year_results: list) -> dict[str, list[str]]:
    """Write one hourly-dispatch sheet per simulated year and return, per metric,
    the list of per-year aggregate cell references (for the Energy sheet to average).

    Each sheet carries the raw hourly series (generation, sales, purchases, PPA
    delivery, penalty, price) and an annual aggregate block that *sums* those
    rows: so the totals feeding the financial model are auditable roll-ups of
    the hourly data, computed live by Excel formulas."""
    import numpy as np  # noqa: F401  (kept for parity / future use)

    refs: dict[str, list[str]] = {k: [] for k in _AGG_ROWS if k != "scale"}

    for idx, res in enumerate(year_results, start=1):
        d = res.dispatch
        prices = res.market_prices
        first_year = getattr(res.scenario, "first_sim_year", 0)
        year_label = (first_year + idx - 1) if first_year else idx
        sheet = wb.create_sheet(f"Hourly Y{idx}")
        sheet.column_dimensions["A"].width = 18

        # Series → arrays
        wind = d.wind_gen.to_numpy()
        pv = d.pv_gen.to_numpy()
        bess_dis = d.bess_dispatch.to_numpy()
        bess_chg = d.bess_store.to_numpy()
        buy = d.market_buy.to_numpy()
        sell = d.market_sell.to_numpy()
        ppa = d.ppa_delivery.to_numpy()
        pen = d.penalty_gen.to_numpy()
        prices.to_numpy()
        total = wind + pv + bess_dis
        index = d.wind_gen.index
        hours = index.hour
        n_hours = len(wind)
        last = _HOURLY_DATA_START + n_hours - 1

        # ── Header / aggregate block ──────────────────────────────────────────
        sheet["A1"] = f"Hourly dispatch: Year {idx} ({year_label})"
        sheet["A1"].font = _TITLE
        sheet["A2"] = (
            "Annual totals below are sums of the hourly rows (× annualisation factor)."
        )
        sheet["A2"].font = Font(italic=True, color="808080")

        def agg(
            key: str, label: str, formula, unit: str, fmt: str = "#,##0.00"
        ) -> None:
            r = _AGG_ROWS[key]
            sheet.cell(r, 1, label).font = Font(bold=True)
            c = sheet.cell(r, 3, formula)
            c.number_format = fmt
            c.fill = _PREFILL
            sheet.cell(r, 4, unit).font = Font(color="808080")

        def rng(col):
            return f"${col}${_HOURLY_DATA_START}:${col}${last}"

        hour_rng = rng(_C_HOUR)
        scale_cell = f"$C${_AGG_ROWS['scale']}"
        solar = f'SUMIFS({{r}},{hour_rng},">=9",{hour_rng},"<17")'

        agg(
            "scale",
            "Annualisation factor (8760 / hours)",
            round(8760.0 / n_hours, 6),
            "×",
            "0.0000",
        )
        agg(
            "ppa_gwh",
            "PPA delivered",
            f"=SUM({rng(_C_PPA)})*{scale_cell}/1000",
            "GWh p.a.",
        )
        agg(
            "excess_solar_gwh",
            "Excess sold: solar hours",
            f"={solar.format(r=rng(_C_SELL))}*{scale_cell}/1000",
            "GWh p.a.",
        )
        agg(
            "excess_nonsolar_gwh",
            "Excess sold: non-solar hours",
            f"=(SUM({rng(_C_SELL)})-{solar.format(r=rng(_C_SELL))})*{scale_cell}/1000",
            "GWh p.a.",
        )
        agg(
            "penalty_gwh",
            "Penalty (undelivered)",
            f"=SUM({rng(_C_PEN)})*{scale_cell}/1000",
            "GWh p.a.",
        )
        agg(
            "total_solar_gwh",
            "Total generation: solar hours",
            f"={solar.format(r=rng(_C_TOTAL))}*{scale_cell}/1000",
            "GWh p.a.",
        )
        agg(
            "total_nonsolar_gwh",
            "Total generation: non-solar hours",
            f"=(SUM({rng(_C_TOTAL)})-{solar.format(r=rng(_C_TOTAL))})*{scale_cell}/1000",
            "GWh p.a.",
        )
        # Volume-weighted capture prices (solar / non-solar hours), guarded for /0
        solar_w = f"({hour_rng}>=9)*({hour_rng}<17)"
        nonsolar_w = f"(({hour_rng}<9)+({hour_rng}>=17))"
        agg(
            "sell_solar_price",
            "Merchant capture: solar hours",
            f"=IFERROR(SUMPRODUCT({solar_w}*{rng(_C_SELL)}*{rng(_C_PRICE)})"
            f"/SUMPRODUCT({solar_w}*{rng(_C_SELL)}),0)",
            "€/MWh",
        )
        agg(
            "sell_nonsolar_price",
            "Merchant capture: non-solar hours",
            f"=IFERROR(SUMPRODUCT({nonsolar_w}*{rng(_C_SELL)}*{rng(_C_PRICE)})"
            f"/SUMPRODUCT({nonsolar_w}*{rng(_C_SELL)}),0)",
            "€/MWh",
        )
        agg(
            "purchase_price",
            "Market purchase price",
            f"=IFERROR(SUMPRODUCT({rng(_C_BUY)}*{rng(_C_PRICE)})/SUM({rng(_C_BUY)}),0)",
            "€/MWh",
        )
        agg(
            "marketbuy_gwh",
            "Market purchase volume",
            f"=SUM({rng(_C_BUY)})*{scale_cell}/1000",
            "GWh p.a.",
        )

        # ── Hourly data ───────────────────────────────────────────────────────
        sheet.cell(_HOURLY_HEADER_ROW, 1)  # ensure header row exists before append
        for j, name in enumerate(_HOURLY_COLS, start=1):
            cell = sheet.cell(_HOURLY_HEADER_ROW, j, name)
            cell.font = _HEADER
            cell.fill = _HEADER_FILL
        # Force appends to begin immediately after the header row.
        sheet._current_row = _HOURLY_HEADER_ROW
        for i in range(n_hours):
            sheet.append(
                [
                    index[i].strftime("%Y-%m-%d %H:%M"),
                    int(hours[i]),
                    round(float(wind[i]), 3),
                    round(float(pv[i]), 3),
                    round(float(bess_dis[i]), 3),
                    round(float(bess_chg[i]), 3),
                    round(float(total[i]), 3),
                    round(float(buy[i]), 3),
                    round(float(sell[i]), 3),
                    round(float(ppa[i]), 3),
                    round(float(pen[i]), 3),
                    round(float(prices[i]), 3),
                ]
            )
        sheet.freeze_panes = f"A{_HOURLY_DATA_START}"

        for k in refs:
            refs[k].append(f"'Hourly Y{idx}'!$C${_AGG_ROWS[k]}")

    return refs


# ── Energy sheet (PyPSA interface) ───────────────────────────────────────────


def _write_energy(
    wb: Workbook,
    e: EnergyInputs,
    hourly_refs: dict[str, list[str]] | None = None,
) -> dict[str, str]:
    ws = wb.create_sheet("Energy")
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    ws["B1"] = "PyPSA Energy Model Results"
    ws["B1"].font = _TITLE
    ws["B2"] = f"Scenario: {e.name}. " + (
        "Annual totals are the average of the per-year sums on the Hourly sheets."
        if hourly_refs
        else "Pre-filled from the energy model."
    )
    ws["B2"].font = Font(italic=True, color="808080")

    cells: dict[str, str] = {}
    row = 4

    def field(label: str, key: str, value, unit: str = "") -> None:
        nonlocal row
        ws.cell(row, 2, label)
        # Summable metrics roll up from the hourly sheets when available.
        if hourly_refs and key in hourly_refs:
            vc = ws.cell(row, 3, f"=AVERAGE({','.join(hourly_refs[key])})")
        else:
            vc = ws.cell(row, 3, value)
        vc.fill = _PREFILL
        vc.border = _BORDER
        if isinstance(value, float):
            vc.number_format = "#,##0.00"
        ws.cell(row, 4, unit).font = Font(color="808080")
        cells[key] = f"Energy!$C${row}"
        row += 1

    field("Onshore wind capacity", "onsw_mw", e.onsw_mw, "MW")
    field("Solar PV capacity", "pv_mw", e.pv_mw, "MW")
    field("BESS power", "bess_mw", e.bess_mw, "MW")
    field("BESS energy", "bess_mwh", e.bess_mwh, "MWh")
    field("Offtaker load", "load_mw", e.load_mw, "MW")
    row += 1
    field("PPA delivered", "ppa_gwh", e.ppa_gwh, "GWh p.a.")
    field(
        "Excess sold: solar hours", "excess_solar_gwh", e.excess_solar_gwh, "GWh p.a."
    )
    field(
        "Excess sold: non-solar hours",
        "excess_nonsolar_gwh",
        e.excess_nonsolar_gwh,
        "GWh p.a.",
    )
    field("Penalty (undelivered)", "penalty_gwh", e.penalty_gwh, "GWh p.a.")
    row += 1
    field(
        "Total generation: solar hours",
        "total_solar_gwh",
        e.total_solar_gwh,
        "GWh p.a.",
    )
    field(
        "Total generation: non-solar hours",
        "total_nonsolar_gwh",
        e.total_nonsolar_gwh,
        "GWh p.a.",
    )
    row += 1
    field(
        "Merchant capture: solar hours", "sell_solar_price", e.sell_solar_price, "€/MWh"
    )
    field(
        "Merchant capture: non-solar hours",
        "sell_nonsolar_price",
        e.sell_nonsolar_price,
        "€/MWh",
    )
    field("Market purchase price", "purchase_price", e.purchase_price, "€/MWh")
    field("Market purchase volume", "marketbuy_gwh", e.marketbuy_gwh, "GWh p.a.")

    return cells


# ── Model sheet (transposed; live formulas) ──────────────────────────────────


# _write_model(wb, p, e, result, inputs_cells, energy_cells)
def _write_model(
    wb: Workbook,
    p: ProjectFinanceInputs,
    e: EnergyInputs,
    result: ProjectFinanceResult,
    IC: dict[str, str],
    EC: dict[str, str],
) -> None:
    ws = wb.create_sheet("Model")
    n = p.model_duration
    tl = _build_timeline(p)
    sc = result.schedule

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 8

    ws["B1"] = "Annual Project-Finance Model"
    ws["B1"].font = _TITLE
    ws["B2"] = (
        "Revenue, opex, depreciation, tax and cash flows are live formulas. "
        "Debt drawdown/IDC and tranche split are toolkit-sized values (green)."
    )
    ws["B2"].font = Font(italic=True, color="808080")

    # Period header
    hdr = 4
    ws.cell(hdr, 2, "Period").font = _HEADER
    ws.cell(hdr, 2).fill = _HEADER_FILL
    ws.cell(hdr, 3, "Unit").font = _HEADER
    ws.cell(hdr, 3).fill = _HEADER_FILL
    for period in range(1, n + 1):
        c = ws.cell(hdr, _pcol(period), period)
        c.font = _HEADER
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    R: dict[str, int] = {}
    row = hdr + 1

    def col(period: int) -> str:
        return get_column_letter(_pcol(period))

    def label_row(name: str, label: str, unit: str = "", section: bool = False) -> int:
        nonlocal row
        r = row
        cell = ws.cell(r, 2, label)
        if section:
            cell.font = _SECTION
            for cc in range(2, _pcol(n) + 1):
                ws.cell(r, cc).fill = _SECTION_FILL
        ws.cell(r, 3, unit).font = Font(color="808080", size=9)
        R[name] = r
        row += 1
        return r

    def put_formula(
        name: str, fn, fmt: str = "#,##0.0", value_fill: bool = False
    ) -> None:
        r = R[name]
        for period in range(1, n + 1):
            cell = ws.cell(r, _pcol(period), fn(period, col(period)))
            cell.number_format = fmt
            if value_fill:
                cell.fill = _PREFILL

    def put_values(
        name: str, arr, fmt: str = "#,##0.0", value_fill: bool = True
    ) -> None:
        r = R[name]
        for period in range(1, n + 1):
            cell = ws.cell(r, _pcol(period), round(float(arr[period - 1]), 6))
            cell.number_format = fmt
            if value_fill:
                cell.fill = _PREFILL

    # ── Flags (values) ───────────────────────────────────────────────────────
    label_row("flags", "Flags", section=True)
    label_row("ops_flag", "Operations flag", "0/1")
    put_values("ops_flag", sc["ops_flag"], "0")
    label_row("ppa_flag", "PPA flag", "0/1")
    put_values("ppa_flag", sc["ppa_flag"], "0")
    nonppa = sc["ops_flag"] - sc["ppa_flag"]
    label_row("nonppa_flag", "Post-PPA flag", "0/1")
    put_values("nonppa_flag", nonppa, "0")
    debt_flag = (
        (result.periods >= tl.ops_start) & (result.periods <= tl.debt_end)
    ).astype(float)
    label_row("debt_flag", "Debt repayment flag", "0/1")
    put_values("debt_flag", debt_flag, "0")

    # ── Indexation (formulas) ─────────────────────────────────────────────────
    label_row("index", "Indexation multiples", section=True)
    label_row("cost_idx", "Cost inflation", "×")
    put_formula(
        "cost_idx",
        lambda pr,
        cl: f"=(1+{IC['cost_inflation']})^({cl}${hdr}+{IC['indexation_offset_years']}-1)",
        "0.000",
    )
    label_row("ppa_idx", "PPA & LGC", "×")
    put_formula(
        "ppa_idx",
        lambda pr,
        cl: f"=(1+{IC['ppa_indexation']})^({cl}${hdr}+{IC['indexation_offset_years']}-1)",
        "0.000",
    )
    label_row("solar_idx", "Solar-hour price", "×")
    put_formula(
        "solar_idx",
        lambda pr,
        cl: f"=(1+{IC['solar_price_inflation']})^({cl}${hdr}+{IC['indexation_offset_years']}-1)",
        "0.000",
    )
    label_row("nonsolar_idx", "Non-solar-hour price", "×")
    put_formula(
        "nonsolar_idx",
        lambda pr,
        cl: f"=(1+{IC['nonsolar_price_inflation']})^({cl}${hdr}+{IC['indexation_offset_years']}-1)",
        "0.000",
    )

    # ── Capital spend (live: cost inputs × capacity × indexation) ─────────────
    # Capex is spread evenly over each tech's construction window (timing
    # fractions baked in per period); devex is the expected cost of reaching
    # FID and is paid as a single lump sum, 100% equity-funded, in the FID
    # period. Cost rates and capacities are live cell references throughout.
    def _fracs(first: int, last: int) -> list[float]:
        arr = [0.0] * n
        if last >= first:
            per = 1.0 / (last - first + 1)
            for pp in range(first, last + 1):
                arr[pp - 1] = per
        return arr

    onsw_con_f = _fracs(*tl.tech_constr(p.onsw_constr_years))
    pv_con_f = _fracs(*tl.tech_constr(p.pv_constr_years))
    bess_con_f = _fracs(*tl.tech_constr(p.bess_constr_years))

    def _devex_tech_fn(rate_key: str, cap_key: str):
        def fn(pr: int, cl: str) -> str:
            return (
                f"=IF({cl}${hdr}={IC['fid_period']},1,0)*{cl}{R['cost_idx']}"
                f"*{IC[rate_key]}*{EC[cap_key]}"
            )

        return fn

    def _capex_tech_fn(build_key: str, conn_key: str, cap_key: str, fracs: list[float]):
        def fn(pr: int, cl: str) -> str:
            return (
                f"={cl}{R['cost_idx']}*{fracs[pr - 1]}"
                f"*({IC[build_key]}+{IC[conn_key]})*{EC[cap_key]}"
            )

        return fn

    label_row("capital", "Capital spend", "€M", section=True)
    label_row("devex_onsw", "Devex: Wind", "€M")
    put_formula("devex_onsw", _devex_tech_fn("onsw_devex", "onsw_mw"))
    label_row("devex_pv", "Devex: Solar PV", "€M")
    put_formula("devex_pv", _devex_tech_fn("pv_devex", "pv_mw"))
    label_row("devex_bess", "Devex: BESS", "€M")
    put_formula("devex_bess", _devex_tech_fn("bess_devex", "bess_mwh"))
    label_row("devex", "Devex: total", "€M")
    put_formula(
        "devex",
        lambda pr,
        cl: f"={cl}{R['devex_onsw']}+{cl}{R['devex_pv']}+{cl}{R['devex_bess']}",
    )
    label_row("capex_onsw", "Capex: Wind", "€M")
    put_formula(
        "capex_onsw",
        _capex_tech_fn(
            "onsw_build_cost", "onsw_connection_cost", "onsw_mw", onsw_con_f
        ),
    )
    label_row("capex_pv", "Capex: Solar PV", "€M")
    put_formula(
        "capex_pv",
        _capex_tech_fn("pv_build_cost", "pv_connection_cost", "pv_mw", pv_con_f),
    )
    label_row("capex_bess", "Capex: BESS", "€M")
    put_formula(
        "capex_bess",
        _capex_tech_fn(
            "bess_build_cost", "bess_connection_cost", "bess_mwh", bess_con_f
        ),
    )
    label_row("capex", "Capex: total", "€M")
    put_formula(
        "capex",
        lambda pr,
        cl: f"={cl}{R['capex_onsw']}+{cl}{R['capex_pv']}+{cl}{R['capex_bess']}",
    )
    label_row("capital_spend", "Total capital spend", "€M")
    put_formula("capital_spend", lambda pr, cl: f"={cl}{R['devex']}+{cl}{R['capex']}")

    # ── Revenue (formulas) ────────────────────────────────────────────────────
    label_row("revenue", "Revenue", "€M", section=True)
    # Volumes by type (GWh p.a.; PPA-period and post-PPA merchant volumes are
    # mutually exclusive in time so each line is non-zero in only one regime).
    label_row("vol_ppa", "PPA volume", "GWh")
    put_formula("vol_ppa", lambda pr, cl: f"={cl}{R['ppa_flag']}*{EC['ppa_gwh']}")
    label_row("vol_penalty", "Penalty volume", "GWh")
    put_formula(
        "vol_penalty", lambda pr, cl: f"={cl}{R['ppa_flag']}*{EC['penalty_gwh']}"
    )
    label_row("vol_merch_solar", "Merchant solar volume", "GWh")
    put_formula(
        "vol_merch_solar",
        lambda pr, cl: (
            f"={cl}{R['ppa_flag']}*{EC['excess_solar_gwh']}+{cl}{R['nonppa_flag']}*{EC['total_solar_gwh']}"
        ),
    )
    label_row("vol_merch_nonsolar", "Merchant non-solar volume", "GWh")
    put_formula(
        "vol_merch_nonsolar",
        lambda pr, cl: (
            f"={cl}{R['ppa_flag']}*{EC['excess_nonsolar_gwh']}+{cl}{R['nonppa_flag']}*{EC['total_nonsolar_gwh']}"
        ),
    )
    label_row("vol_lgc", "LGC volume", "GWh")
    put_formula(
        "vol_lgc",
        lambda pr, cl: f"={cl}{R['vol_merch_solar']}+{cl}{R['vol_merch_nonsolar']}",
    )

    # Prices by type (€/MWh). Merchant prices are escalated only if the energy
    # inputs aren't already year-specific (otherwise price growth would be
    # double-counted).
    esc = p.escalate_merchant_prices
    label_row("price_ppa", "PPA price", "€/MWh")
    put_formula("price_ppa", lambda pr, cl: f"={IC['ppa_tariff']}*{cl}{R['ppa_idx']}")
    label_row("price_penalty", "Penalty price", "€/MWh")
    put_formula(
        "price_penalty",
        lambda pr,
        cl: f"={IC['ppa_tariff']}*{IC['penalty_multiple']}*{cl}{R['ppa_idx']}",
    )
    label_row("price_merch_solar", "Merchant solar price", "€/MWh")
    put_formula(
        "price_merch_solar",
        lambda pr, cl: (
            f"={EC['sell_solar_price']}{(f'*{cl}' + str(R['solar_idx'])) if esc else ''}"
        ),
    )
    label_row("price_merch_nonsolar", "Merchant non-solar price", "€/MWh")
    put_formula(
        "price_merch_nonsolar",
        lambda pr, cl: (
            f"={EC['sell_nonsolar_price']}{(f'*{cl}' + str(R['nonsolar_idx'])) if esc else ''}"
        ),
    )
    label_row("price_lgc", "LGC price", "€/MWh")
    put_formula("price_lgc", lambda pr, cl: f"={IC['lgc_price']}*{cl}{R['ppa_idx']}")

    # Revenue by type = volume (GWh) × price (€/MWh) / 1000 = €M, summing to
    # total revenue.
    label_row("rev_ppa", "PPA revenue", "€M")
    put_formula(
        "rev_ppa", lambda pr, cl: f"={cl}{R['vol_ppa']}*{cl}{R['price_ppa']}/1000"
    )
    label_row("cost_penalty", "Penalty cost", "€M")
    put_formula(
        "cost_penalty",
        lambda pr, cl: f"={cl}{R['vol_penalty']}*{cl}{R['price_penalty']}/1000",
    )
    label_row("rev_merch_solar", "Merchant solar revenue", "€M")
    put_formula(
        "rev_merch_solar",
        lambda pr, cl: f"={cl}{R['vol_merch_solar']}*{cl}{R['price_merch_solar']}/1000",
    )
    label_row("rev_merch_nonsolar", "Merchant non-solar revenue", "€M")
    put_formula(
        "rev_merch_nonsolar",
        lambda pr,
        cl: f"={cl}{R['vol_merch_nonsolar']}*{cl}{R['price_merch_nonsolar']}/1000",
    )
    label_row("rev_lgc", "LGC / GO revenue", "€M")
    put_formula(
        "rev_lgc", lambda pr, cl: f"={cl}{R['vol_lgc']}*{cl}{R['price_lgc']}/1000"
    )
    label_row("net_contracted", "Net contracted revenue", "€M")
    put_formula(
        "net_contracted", lambda pr, cl: f"={cl}{R['rev_ppa']}-{cl}{R['cost_penalty']}"
    )
    label_row("net_uncontracted", "Net uncontracted revenue", "€M")
    put_formula(
        "net_uncontracted",
        lambda pr,
        cl: f"={cl}{R['rev_merch_solar']}+{cl}{R['rev_merch_nonsolar']}+{cl}{R['rev_lgc']}",
    )
    label_row("total_rev", "Total revenue", "€M")
    put_formula(
        "total_rev",
        lambda pr, cl: f"={cl}{R['net_contracted']}+{cl}{R['net_uncontracted']}",
    )

    # ── Opex / EBITDA (formulas) ──────────────────────────────────────────────
    label_row("opex_sec", "Operating costs", "€M", section=True)
    fixed_om_expr = f"({IC['onsw_fixed_om']}*{EC['onsw_mw']}+{IC['pv_fixed_om']}*{EC['pv_mw']}+{IC['bess_fixed_om']}*{EC['bess_mwh']})"
    label_row("opex", "Total O&M expenses", "€M")
    put_formula(
        "opex",
        lambda pr,
        cl: f"={cl}{R['ops_flag']}*{fixed_om_expr}+{IC['ancillary_pct']}*{cl}{R['total_rev']}",
    )
    label_row("ebitda", "EBITDA", "€M")
    put_formula("ebitda", lambda pr, cl: f"={cl}{R['total_rev']}-{cl}{R['opex']}")

    # ── Debt (toolkit values + interest formula reference) ────────────────────
    label_row("debt", "Debt schedule", "€M", section=True)
    label_row("debt_draw", "Debt drawdown", "€M")
    put_values("debt_draw", sc["debt_draw"])
    label_row("idc", "Interest during construction", "€M")
    put_values("idc", sc["idc"])
    label_row("interest", "Term loan interest", "€M")
    put_values("interest", sc["interest"])
    label_row("loan_repay", "Loan repayment", "€M")
    put_values("loan_repay", sc["loan_repay"])

    # ── Depreciation (live, straight-line capped at the asset base) ───────────
    firstcol, lastcol = col(1), col(n)
    label_row("dep", "Depreciation", "€M", section=True)

    # Asset bases (live): tax = capex only; book = devex + capex + capitalised IDC.
    label_row("tax_base", "Tax asset base", "€M")
    ws.cell(
        R["tax_base"], _pcol(1), f"=SUM({firstcol}{R['capex']}:{lastcol}{R['capex']})"
    ).number_format = "#,##0.0"
    label_row("book_base", "Book asset base", "€M")
    ws.cell(
        R["book_base"],
        _pcol(1),
        (
            f"=SUM({firstcol}{R['devex']}:{lastcol}{R['devex']})"
            f"+SUM({firstcol}{R['capex']}:{lastcol}{R['capex']})"
            f"+SUM({firstcol}{R['idc']}:{lastcol}{R['idc']})"
        ),
    ).number_format = "#,##0.0"

    def _dep_fn(self_row: int, base_row: int, rate_cell: str):
        # Straight-line at `rate` on the asset base, but never depreciate more
        # than the remaining book value (cumulative prior depreciation in-row).
        base = f"${firstcol}${base_row}"

        def fn(pr: int, cl: str) -> str:
            prior = (
                "0"
                if pr == 1
                else f"SUM(${firstcol}{self_row}:{col(pr - 1)}{self_row})"
            )
            return f"={cl}{R['ops_flag']}*MIN({base}*{rate_cell},MAX({base}-{prior},0))"

        return fn

    label_row("tax_dep", "Tax depreciation", "€M")
    put_formula(
        "tax_dep", _dep_fn(R["tax_dep"], R["tax_base"], IC["tax_depreciation_rate"])
    )
    label_row("book_dep", "Book depreciation", "€M")
    put_formula(
        "book_dep", _dep_fn(R["book_dep"], R["book_base"], IC["book_depreciation_rate"])
    )

    # ── P&L tax (live, with loss carry-forward) ───────────────────────────────
    label_row("pl", "Profit & tax", "€M", section=True)
    label_row("pbt", "Profit before tax", "€M")
    put_formula(
        "pbt",
        lambda pr, cl: f"={cl}{R['ebitda']}-{cl}{R['interest']}-{cl}{R['book_dep']}",
    )
    label_row("taxable", "Taxable income", "€M")
    put_formula(
        "taxable",
        lambda pr, cl: (
            f"={cl}{R['ebitda']}-{cl}{R['interest']}-{cl}{R['tax_dep']}-{cl}{R['devex']}"
        ),
    )
    label_row("carry", "Carry-forward losses", "€M")
    put_formula(
        "carry",
        lambda pr, cl: (
            f"=MIN(0,{cl}{R['taxable']})"
            if pr == 1
            else f"=MIN(0,{cl}{R['taxable']}+{col(pr - 1)}{R['carry']})"
        ),
    )
    label_row("tax", "Income tax", "€M")
    put_formula(
        "tax",
        lambda pr, cl: (
            f"=MAX(0,{cl}{R['taxable']})*{IC['corp_tax_rate']}"
            if pr == 1
            else f"=MAX(0,{cl}{R['taxable']}+{col(pr - 1)}{R['carry']})*{IC['corp_tax_rate']}"
        ),
    )
    label_row("pat", "Profit after tax", "€M")
    put_formula("pat", lambda pr, cl: f"={cl}{R['pbt']}-{cl}{R['tax']}")

    # ── Returns (formulas) ────────────────────────────────────────────────────
    label_row("returns", "Returns", "€M", section=True)
    label_row("fcff", "FCFF (project)", "€M")
    put_formula(
        "fcff",
        lambda pr, cl: (
            f"={cl}{R['ops_flag']}*({cl}{R['ebitda']}-{cl}{R['tax']})-{cl}{R['capital_spend']}"
        ),
    )
    label_row("equity_spend", "Equity investment", "€M")
    put_formula(
        "equity_spend", lambda pr, cl: f"={cl}{R['capital_spend']}-{cl}{R['debt_draw']}"
    )
    label_row("fcfe", "FCFE (equity)", "€M")
    put_formula(
        "fcfe",
        lambda pr, cl: (
            f"={cl}{R['ops_flag']}*({cl}{R['pat']}+{cl}{R['book_dep']}-{cl}{R['loan_repay']})-{cl}{R['equity_spend']}"
        ),
    )
    label_row("cfads", "CFADS", "€M")
    put_formula(
        "cfads",
        lambda pr, cl: f"={cl}{R['ops_flag']}*({cl}{R['ebitda']}-{cl}{R['tax']})",
    )
    label_row("dscr", "DSCR", "ratio")
    put_formula(
        "dscr",
        lambda pr, cl: (
            f"=IF(({cl}{R['interest']}+{cl}{R['loan_repay']})>0,"
            f'{cl}{R["cfads"]}/({cl}{R["interest"]}+{cl}{R["loan_repay"]}),"")'
        ),
        "0.00",
    )

    # ── Cash flow statement (formulas) ────────────────────────────────────────
    label_row("cf", "Cash flow statement", "€M", section=True)
    label_row("cfo", "Cash from operations", "€M")
    put_formula(
        "cfo",
        lambda pr, cl: (
            f"={cl}{R['ops_flag']}*({cl}{R['ebitda']}-{cl}{R['interest']}-{cl}{R['tax']})"
        ),
    )
    label_row("cfi", "Cash from investing", "€M")
    put_formula("cfi", lambda pr, cl: f"=-{cl}{R['capital_spend']}")
    label_row("cff", "Cash from financing", "€M")
    put_formula(
        "cff",
        lambda pr,
        cl: f"={cl}{R['debt_draw']}+{cl}{R['equity_spend']}-{cl}{R['loan_repay']}",
    )
    label_row("net_cash_flow", "Net cash flow", "€M")
    put_formula(
        "net_cash_flow", lambda pr, cl: f"={cl}{R['cfo']}+{cl}{R['cfi']}+{cl}{R['cff']}"
    )

    # ── Balance sheet (formulas, running roll-forwards) ───────────────────────
    # No dividends/distributions are modelled yet: all profit and cash
    # generated is retained, so retained earnings and cash simply accumulate.
    def _running(name: str, period_expr):
        def fn(pr: int, cl: str) -> str:
            expr = period_expr(pr, cl)
            return expr if pr == 1 else f"={col(pr - 1)}{R[name]}+{expr.lstrip('=')}"

        return fn

    label_row("bs", "Balance sheet", "€M", section=True)
    label_row("cash_balance", "Cash balance", "€M")
    put_formula(
        "cash_balance",
        _running("cash_balance", lambda pr, cl: f"={cl}{R['net_cash_flow']}"),
    )
    label_row("ppe_net", "PP&E, net", "€M")
    put_formula(
        "ppe_net",
        _running(
            "ppe_net",
            lambda pr,
            cl: f"={cl}{R['capex']}+{cl}{R['devex']}+{cl}{R['idc']}-{cl}{R['book_dep']}",
        ),
    )
    label_row("debt_balance", "Debt balance", "€M")
    put_formula(
        "debt_balance",
        _running(
            "debt_balance",
            lambda pr,
            cl: f"={cl}{R['debt_draw']}+{cl}{R['idc']}-{cl}{R['loan_repay']}",
        ),
    )
    label_row("share_capital", "Share capital", "€M")
    put_formula(
        "share_capital",
        _running("share_capital", lambda pr, cl: f"={cl}{R['equity_spend']}"),
    )
    label_row("retained_earnings", "Retained earnings", "€M")
    put_formula(
        "retained_earnings",
        _running("retained_earnings", lambda pr, cl: f"={cl}{R['pat']}"),
    )
    label_row("total_assets", "Total assets", "€M")
    put_formula(
        "total_assets", lambda pr, cl: f"={cl}{R['ppe_net']}+{cl}{R['cash_balance']}"
    )
    label_row("total_liabilities", "Total liabilities", "€M")
    put_formula("total_liabilities", lambda pr, cl: f"={cl}{R['debt_balance']}")
    label_row("total_equity_bs", "Total equity", "€M")
    put_formula(
        "total_equity_bs",
        lambda pr, cl: f"={cl}{R['share_capital']}+{cl}{R['retained_earnings']}",
    )
    label_row("bs_check", "Balance check (Assets − Liab. − Equity)", "€M")
    put_formula(
        "bs_check",
        lambda pr, cl: (
            f"={cl}{R['total_assets']}-{cl}{R['total_liabilities']}-{cl}{R['total_equity_bs']}"
        ),
        "0.0000",
    )

    # remember key ranges for the Outputs sheet
    first, last = _pcol(1), _pcol(n)
    fl, ll = get_column_letter(first), get_column_letter(last)
    wb._fm_ranges = {  # type: ignorEC[attr-defined]
        "fcff": f"Model!{fl}{R['fcff']}:{ll}{R['fcff']}",
        "fcfe": f"Model!{fl}{R['fcfe']}:{ll}{R['fcfe']}",
        "ebitda": f"Model!{fl}{R['ebitda']}:{ll}{R['ebitda']}",
        "dscr": f"Model!{fl}{R['dscr']}:{ll}{R['dscr']}",
        "bs_check": f"Model!{fl}{R['bs_check']}:{ll}{R['bs_check']}",
    }


# ── Outputs sheet ─────────────────────────────────────────────────────────────


def _write_outputs(wb: Workbook, result: ProjectFinanceResult) -> None:
    ws = wb.create_sheet("Outputs", 0)  # first sheet
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    rng = getattr(wb, "_fm_ranges", {})

    ws["B1"] = "Financial Model: Key Outputs"
    ws["B1"].font = _TITLE
    ws["B2"] = f"Scenario: {result.energy.name}"
    ws["B2"].font = Font(italic=True, color="808080")

    row = 4

    def kpi(label: str, value, fmt: str, formula: str | None = None) -> None:
        nonlocal row
        ws.cell(row, 2, label).font = Font(bold=True)
        c = ws.cell(row, 3)
        c.value = formula if formula is not None else value
        c.number_format = fmt
        c.fill = _PREFILL
        c.border = _BORDER
        row += 1

    kpi(
        "Project IRR (FCFF)",
        result.project_irr,
        "0.0%",
        f'=IFERROR(IRR({rng.get("fcff", "")}),"n/a")' if rng.get("fcff") else None,
    )
    kpi(
        "Equity IRR (FCFE)",
        result.equity_irr,
        "0.0%",
        f'=IFERROR(IRR({rng.get("fcfe", "")}),"n/a")' if rng.get("fcfe") else None,
    )
    kpi("NPV @ WACC (project)", result.npv_project, '#,##0.0 "€M"')
    kpi("Gearing", result.gearing, "0.0%")
    kpi("Total funding (incl. IDC)", result.total_capex, '#,##0.0 "€M"')
    kpi("Total debt", result.total_debt, '#,##0.0 "€M"')
    kpi("Total equity", result.total_equity, '#,##0.0 "€M"')
    kpi("Minimum DSCR", result.min_dscr, "0.00")
    kpi("Average DSCR", result.avg_dscr, "0.00")
    kpi("Equity payback", result.payback_years, '0.0 "yrs"')
    kpi(
        "Max balance-sheet check (should be ~0)",
        result.max_bs_check,
        '#,##0.0000 "€M"',
        f"=MAX(MAX({rng.get('bs_check', '')}),-MIN({rng.get('bs_check', '')}))"
        if rng.get("bs_check")
        else None,
    )
    kpi("LCOE", result.lcoe, '#,##0.0 "€/MWh"')

    ws.cell(
        row + 1,
        2,
        "IRRs and the capex→depreciation→tax→cash-flow chain recompute live from the "
        "Model sheet. Debt sizing/IDC are pre-solved (circular); re-run the toolkit to "
        "re-size debt after large cost changes.",
    ).font = Font(italic=True, color="A0A0A0", size=9)


# ── Notes sheet ───────────────────────────────────────────────────────────────


def _write_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Notes")
    ws.column_dimensions["B"].width = 100
    ws["B1"] = "Model notes & simplifications"
    ws["B1"].font = _TITLE
    notes = [
        "This workbook is a streamlined export of the PyPSA-PPA toolkit's project-finance model.",
        "",
        "Live (formula-driven, recompute on edit):",
        "  • Hourly sheets (one per simulated year) hold the full hourly dispatch; the",
        "    Energy-sheet annual totals are SUM/SUMIFS roll-ups of those hours, averaged",
        "    across years. Edit the hourly data and the totals (and the model) follow.",
        "  • Capex: per-technology build/connection cost × capacity × indexation, spread",
        "    over each tech's construction window (spend timing baked per period).",
        "  • Devex: per-technology expected cost of reaching FID, paid as a single lump",
        "    sum in the FID period, 100% equity-funded (never refinanced by debt).",
        "  • Indexation multipliers; revenue by type (volume × price), opex, EBITDA.",
        "  • Book/tax depreciation (straight-line, capped at the live asset base). Devex",
        "    is capitalised into the book base but fully expensed for tax when paid.",
        "  • Taxable income, loss carry-forward and income tax.",
        "  • PBT, PAT, FCFF, FCFE, DSCR, and the Project/Equity IRR outputs.",
        "  • Cash flow statement (operating / investing / financing) and balance sheet",
        "    (PP&E, cash, debt, share capital, retained earnings), with a balance check",
        "    row (Assets − Liabilities − Equity) that should read ~0 in every period.",
        "",
        "Toolkit-sized values (green): edit to override:",
        "  • Debt drawdown and IDC are circular (debt size depends on IDC which depends",
        "    on drawdown), so they are pre-solved. Changing capex therefore updates",
        "    returns but not the debt amount: re-run the toolkit to re-size debt.",
        "",
        "Simplifications (consistent with the source model):",
        "  • No working capital, no dividends/distributions yet, no terminal or",
        "    decommissioning value. Retained earnings and cash simply accumulate.",
        "  • One representative operating year, escalated by indexation.",
        "  • Solar hours defined as 09:00–17:00.",
    ]
    for i, line in enumerate(notes):
        ws.cell(2 + i, 2, line)
